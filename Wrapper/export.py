import json

import numpy as np
from openpyxl import load_workbook

algorithms = ["MEES", "Random", "HEFT_EDF", "Fuzzy"]
tasks_load = "load-3"
task_dominance = "hard"
config = "config-1"

fog_count = 3
edge_count = 10
total_time = 510_000  # millisecond

file_path = "./results.xlsx"

wb = load_workbook(file_path)


def create_sheet(sheet_name: str):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    return ws


def energy_latency():
    # energy consumption & latency results
    energy_trace_cloud: dict = {}
    energy_trace_fog: dict = {}
    energy_trace_edge: dict = {}
    latency_max: dict = {}
    latency_min: dict = {}
    latency_sum: dict = {}
    latency_count: dict = {}

    time_step = 1000

    for algorithm in algorithms:

        energy_trace_cloud[algorithm] = np.array([0.0 for i in range(total_time // time_step)])
        energy_trace_fog[algorithm] = np.array([0.0 for i in range(total_time // time_step)])
        energy_trace_edge[algorithm] = np.array([0.0 for i in range(total_time // time_step)])

        latency_max[algorithm] = np.array([0.0 for i in range(total_time // time_step)])
        latency_min[algorithm] = np.array([0.0 for i in range(total_time // time_step)])
        latency_sum[algorithm] = np.array([0.0 for i in range(total_time // time_step)])
        latency_count[algorithm] = np.array([0.0 for i in range(total_time // time_step)])

        for e in range(edge_count):
            with open(f'Results/{config}/Edge/tasks-{algorithm}-{tasks_load}-{task_dominance}-{e}.json', 'r') as file:
                tasks = json.load(file)

                for task in tasks:
                    arrival_time = task["arrivalTime"]
                    if arrival_time >= total_time:
                        continue

                    execution = sum([ex[2] - ex[1] for ex in task["executionTimes"]])
                    latency = task["latency"]
                    if latency is None:
                        latency = execution
                    else:
                        latency = int(latency + 10 * time_step)  # L base

                    finish_time = arrival_time + latency

                    arrival_time = arrival_time // time_step
                    finish_time = finish_time // time_step
                    latency = latency / time_step

                    for l in range(arrival_time, finish_time):
                        if latency_max[algorithm][l] < latency:
                            latency_max[algorithm][l] = latency
                        if latency_min[algorithm][l] == 0 or latency_min[algorithm][l] > latency:
                            latency_min[algorithm][l] = latency
                        latency_sum[algorithm][l] += latency
                        latency_count[algorithm][l] += 1

        if True:
            with open(f'Results/{config}/Cloud/processors-{algorithm}-{tasks_load}-{task_dominance}.json',
                      'r') as file:
                processors = json.load(file)

            energy_trace_cloud_temp = np.array([processors[0]["idlePower"] for i in range(total_time // time_step)])
            if algorithm == "MEES":
                energy_trace_cloud_temp[:] = 0

            overhead_power = processors[0]["activePower"] - processors[0]["idlePower"]

            for processor in processors:
                for allocation in processor["allocations"]:
                    start = allocation["startTime"]
                    end = allocation["endTime"]
                    if start > total_time:
                        continue
                    if end > total_time:
                        end = total_time

                    energy_trace_cloud_temp[start // time_step: end // time_step] += overhead_power / len(processors)

            energy_trace_cloud[algorithm][:] += energy_trace_cloud_temp

        for i in range(fog_count):
            with open(f'Results/{config}/Fog/processors-{algorithm}-{tasks_load}-{task_dominance}-{i}.json',
                      'r') as file:
                processors = json.load(file)

            energy_trace_fog_temp = np.array([processors[0]["idlePower"] for i in range(total_time // time_step)])
            if algorithm == "MEES":
                energy_trace_fog_temp[:] = 0

            overhead_power = processors[0]["activePower"] - processors[0]["idlePower"]

            for processor in processors:
                for allocation in processor["allocations"]:
                    start = allocation["startTime"]
                    end = allocation["endTime"]
                    if start > total_time:
                        continue
                    if end > total_time:
                        end = total_time

                    energy_trace_fog_temp[start // time_step: end // time_step] += overhead_power / len(processors)

            energy_trace_fog[algorithm][:] += energy_trace_fog_temp

        for i in range(edge_count):
            with open(f'Results/{config}/Edge/processors-{algorithm}-{tasks_load}-{task_dominance}-{i}.json',
                      'r') as file:
                processors = json.load(file)

            energy_trace_edge_temp = np.array([processors[0]["idlePower"] for i in range(total_time // time_step)],
                                              dtype=float)
            overhead_power = processors[0]["activePower"] - processors[0]["idlePower"]

            for processor in processors:
                for allocation in processor["allocations"]:
                    start = allocation["startTime"]
                    end = allocation["endTime"]
                    if start > total_time:
                        continue
                    if end > total_time:
                        end = total_time

                    energy_trace_edge_temp[start // time_step: end // time_step] += overhead_power / len(processors)

            energy_trace_edge[algorithm][:] += energy_trace_edge_temp

    # write in excel
    ws = create_sheet("E - L(Max)")
    row = 1
    for i in range(total_time // time_step):
        ws.cell(row=row, column=2 + i, value=i + 1)
    for algorithm in algorithms:
        row += 1
        ws.cell(row=row, column=1, value=algorithm)
        for i in range(total_time // time_step):
            ws.cell(row=row, column=2 + i, value=energy_trace_edge[algorithm][i]
                                                 + energy_trace_fog[algorithm][i]
                                                 + energy_trace_cloud[algorithm][i])

    row = 12
    for i in range(total_time // time_step):
        ws.cell(row=row, column=2 + i, value=i + 1)
    for algorithm in algorithms:
        row += 1
        ws.cell(row=row, column=1, value=algorithm)
        for i in range(total_time // time_step):
            ws.cell(row=row, column=2 + i, value=latency_max[algorithm][i])

    wb.save(file_path)

    ws = create_sheet("E - L(Min)")
    row = 1
    for i in range(total_time // time_step):
        ws.cell(row=row, column=2 + i, value=i + 1)
    for algorithm in algorithms:
        row += 1
        ws.cell(row=row, column=1, value=algorithm)
        for i in range(total_time // time_step):
            ws.cell(row=row, column=2 + i, value=energy_trace_edge[algorithm][i]
                                                 + energy_trace_fog[algorithm][i]
                                                 + energy_trace_cloud[algorithm][i])

    row = 12
    for i in range(total_time // time_step):
        ws.cell(row=row, column=2 + i, value=i + 1)
    for algorithm in algorithms:
        row += 1
        ws.cell(row=row, column=1, value=algorithm)
        for i in range(total_time // time_step):
            ws.cell(row=row, column=2 + i, value=latency_min[algorithm][i])

    wb.save(file_path)

    ws = create_sheet("E - L(Avg)")
    row = 1
    for i in range(total_time // time_step):
        ws.cell(row=row, column=2 + i, value=i + 1)
    for algorithm in algorithms:
        row += 1
        ws.cell(row=row, column=1, value=algorithm)
        for i in range(total_time // time_step):
            ws.cell(row=row, column=2 + i, value=energy_trace_edge[algorithm][i]
                                                 + energy_trace_fog[algorithm][i]
                                                 + energy_trace_cloud[algorithm][i])

    row = 12
    for i in range(total_time // time_step):
        ws.cell(row=row, column=2 + i, value=i + 1)
    for algorithm in algorithms:
        row += 1
        ws.cell(row=row, column=1, value=algorithm)
        for i in range(total_time // time_step):
            latency = 0
            if latency_sum[algorithm][i] > 0:
                latency = int(latency_sum[algorithm][i] // latency_count[algorithm][i])
            ws.cell(row=row, column=2 + i, value=latency)

    wb.save(file_path)


def makespan_tasks():
    makespans: dict = {}
    tasks_count: dict = {}
    time_step = 1000

    for algorithm in algorithms:

        makespans[algorithm] = np.array([0.0 for i in range(total_time // time_step)])
        tasks_count[algorithm] = np.array([0.0 for i in range(total_time // time_step)])

        for e in range(edge_count):
            with open(f'Results/{config}/Edge/tasks-{algorithm}-{tasks_load}-{task_dominance}-{e}.json', 'r') as file:
                tasks = json.load(file)

                for task in tasks:
                    if task["type"] != 'HARD':
                        continue

                    arrival_time = task["arrivalTime"]
                    if arrival_time >= total_time:
                        continue

                    makespan = task["executionTimes"][-1][2] - arrival_time

                    arrival_time = arrival_time // time_step
                    finish_time = task["executionTimes"][0][2]
                    if finish_time > total_time:
                        finish_time = total_time
                    finish_time = finish_time // time_step

                    for t in range(arrival_time, finish_time):
                        makespans[algorithm][t] += makespan
                        tasks_count[algorithm][t] += 1
    # write in excel
    ws = create_sheet("M - S")
    row = 1
    for i in range(total_time // time_step):
        ws.cell(row=row, column=2 + i, value=i + 1)
    for algorithm in algorithms:
        row += 1
        ws.cell(row=row, column=1, value=algorithm)
        for i in range(total_time // time_step):
            makespan = makespans[algorithm][i]
            if makespan != 0:
                makespan = makespan // tasks_count[algorithm][i]
            ws.cell(row=row, column=2 + i, value=makespan)

    row = 12
    for i in range(total_time // time_step):
        ws.cell(row=row, column=2 + i, value=i + 1)
    for algorithm in algorithms:
        row += 1
        ws.cell(row=row, column=1, value=algorithm)
        for i in range(total_time // time_step):
            ws.cell(row=row, column=2 + i, value=tasks_count[algorithm][i])

    wb.save(file_path)


def success_rate():
    loads = [f"load-{i}" for i in [2, 3, 4]]

    success_count = {}
    total = {}

    for load in loads:
        for algorithm in algorithms:
            success_count[(load, algorithm)] = 0
            total[(load, algorithm)] = 0

            for e in range(edge_count):
                with open(f'Results/{config}/Edge/tasks-{algorithm}-{load}-{task_dominance}-{e}.json',
                          'r') as file:
                    tasks = json.load(file)

                    for task in tasks:
                        if task["type"] != "HARD":
                            continue

                        arrival_time = task["arrivalTime"]
                        deadline = task["deadline"] + task["arrivalTime"]

                        if arrival_time >= total_time:
                            continue

                        if task["executionTimes"][-1][2] < deadline and task["remainingExecutionCost"] == 0:
                            success_count[(load, algorithm)] += 1
                        total[(load, algorithm)] += 1

    # write in excel
    print(total, success_count)
    ws = create_sheet("Success Rate")
    row = 1
    for algorithm in algorithms:
        row += 1
        ws.cell(row=row, column=1, value=algorithm)

        for i in range(len(loads)):
            sr = success_count[(loads[i], algorithm)] / total[(loads[i], algorithm)]
            ws.cell(row=row, column=2 + i, value=sr)

    wb.save(file_path)


if __name__ == '__main__':
    energy_latency()
    makespan_tasks()
    success_rate()
